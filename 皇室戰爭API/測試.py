import openpyxl
import requests
import json
import re
import time
start_time = time.time()
API_KEY = "eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzUxMiIsImtpZCI6IjI4YTMxOGY3LTAwMDAtYTFlYi03ZmExLTJjNzQzM2M2Y2NhNSJ9.eyJpc3MiOiJzdXBlcmNlbGwiLCJhdWQiOiJzdXBlcmNlbGw6Z2FtZWFwaSIsImp0aSI6IjMwMTdiMmNjLTJmZDktNDEyNi05OTFjLWNkYWZjZWY5N2IxNiIsImlhdCI6MTY5NzAyMTgwNiwic3ViIjoiZGV2ZWxvcGVyLzVlYTM1ZjUwLTEwMTMtZjkzMC02OTEyLTIzNTU2M2QwN2FhYyIsInNjb3BlcyI6WyJyb3lhbGUiXSwibGltaXRzIjpbeyJ0aWVyIjoiZGV2ZWxvcGVyL3NpbHZlciIsInR5cGUiOiJ0aHJvdHRsaW5nIn0seyJjaWRycyI6WyI0OS4xNTkuMjQ4LjEzIl0sInR5cGUiOiJjbGllbnQifV19.Qz1nRZK6K8jb9cN2dEbDcnKql-xc1bnDJviAUWuCvFUQehHBIPvZfw9zUUz_5n6VyNVd93kcLuk5BjyAZaXERg"
headers = {
    "Authorization": "Bearer {}".format(API_KEY)
}

player_tags = [
    "%2322R920J00", 
    # Xiang♥
      "%23QV9RU0J", 
    # ༤༒康師父༒༄
      "%2328RC9RGR8", 
    #   倫哥
      "%23PVJJCC82", 
    #   GIF
      "%23PJVYVPRG", 
    #   Point
      "%23VYG80P8G", 
    #  塔兔 
      "%239CLQRCYVP", 
    #   Tžↁ-落筆成墨
      "%23Q2CUV99", 
    #   
      # "%238Q902LVJP", 
    #   The_Hesitation
      "%23RUVRL29GV", 
    #   小錯
      "%232YCYQYVG9", 
    #   戰▪滅☆《冰皇》“巔峰▪時刻"
      "%2322VCCPLQ", 
    #   ✨Honey beef✨
      "%23988RYC0R9", 
    #   手談
      "%239YUCLQC2V", 
    #   名字
      "%23PVP0JR2VU", 
    #   BlueIce
      "%238Y22QPCU0", 
    #   大肚李敏鎬
      "%2328QCYLV2P", 
    #   Justin Hsu
      "%23PQYYLU2G", 
    #   Taoki
      "%23YR2RR2CC", 
    #   高質感妮妮
      "%23LPR2G2CJ",
    #   Bridge
      "%239C9VJLRC8", 
    #   無肉令人瘦♨️
      "%23CPRYQVL8", 
    #   呵呵
      "%23LLYGYY29Q", 
    #   wiake
      "%23RY00UJL29", 
    #   雪色如卿
      "%23YQ908ULQC", 
    #   paul1210
      "%239UVLUYYQ", 
    #   絕戀。柔
      "%239R0VQP9RQ", 
    #   ☢️P.E.K.K.A™️♾♾
      "%238JQGQY0", 
    #   Ice•peanut
      "%23LQGG9UY2P", 
    #   Quincy極光
      "%238YQRVGPQY", 
    #   ᴀɴɢ★ʀᴀᴘʜᴀᴇʟ⚡
      "%23PRLV0QQPJ",
    #    ~DuckYears~
      "%239P8JQ9908", 
    #   [TW]抹茶
      "%23R8JP0JLL", 
    #   阿芳餒
      "%23UR99PYYR", 
    #   PK❤️GY
      "%23RJRR9UJ", 
    #   Boss Benedict
      "%23Y2UP2JY82",  
    #   小心偷塔
      "%23UUQL0RCC",
    #   ATW｜Wizardᵀᴹ
      "%23LV890YUUU", 
    #   一名術士
      "%23JCU9J802", 
    #   ᴀɴɢ★ᴍᴇᴛᴀᴛʀᴏɴ⚡
      "%23P0PPCJL2U",
    #   缱绻｜Lingering
      "%23P9PRPUCR8",
    #   白上フブキ
      "%23PQJLJ00Y0",
    #   MONPAN
      "%23LG9U2JPC",
    #   Jes
      "%2322CG0PCJR",
    #   JaYan
      "%23PCU80CLLP",
    #   장첸밍
      "%238RVQ09RP8",
    #   AceRank
        "%238P929U0G",
    # 神意志帝國
        "%232PLPV892",
    # Mr★NeverDie 
        "%232C08R0P2",
    # ゆき
        "%23L8QCL2U0",
    #《SECRET》 
      ]
              
wb = openpyxl.Workbook()
ws = wb.active

ws["A1"] = "遊戲ID"
ws["B1"] = "最佳賽季"
ws["C1"] = "Rating"
ws["D1"] = "上一賽季"
ws["E1"] = "Rating"
ws["F1"] = "經典挑戰12勝次數"
ws["G1"] = "終極挑戰12勝次數"
ws["H1"] = "遊玩時間"
ws["I1"] = "表情收集次數"
ws["J1"] = "戰旗收集次數"
ws["K1"] = "加入部落次數(?)"
ws["L1"] = "友誼戰獲勝次數"
ws["M1"] = "星光點數"
ws["N1"] = "全部經驗值"

def clean_result(result):
    if result is None:
        return ""
    return result[result.rfind(" ") + 1:].split("}")[0]


def keep_numbers(string):
  
  filtered_string = filter(str.isdigit, string)
  return "".join(filtered_string)

def get_numbers_by_position_and_delete(variable, start_position, end_position):

  numbers = str(variable)[start_position:end_position]
  filtered_numbers = filter(str.isdigit, numbers)
  return "".join(filtered_numbers)

x=0




for player_tag in player_tags:
    response = requests.get(f"https://api.clashroyale.com/v1/players/{player_tag}", headers=headers)
    
    player_data = response.json()
    best = json.dumps(player_data.get("bestPathOfLegendSeasonResult"))
    last = json.dumps(player_data.get("lastPathOfLegendSeasonResult"))

    best2 = clean_result(best)
    last2= clean_result(last)

    best3 = get_numbers_by_position_and_delete(keep_numbers(best), 2, 6)
    last3 = get_numbers_by_position_and_delete(keep_numbers(last), 2, 6)

    y=player_data["tag"]

    Classic12WinsCout =0 
    Grand12WinsCout = 0

    starPoints = player_data["starPoints"]

    totalExpPoints = player_data["totalExpPoints"]
    
    for Classic12Wins in player_data["badges"]:
       if Classic12Wins["name"] == "Classic12Wins":
        Classic12WinsCout = Classic12Wins["progress"]

    for Grand12Wins in player_data["badges"]:
       if Grand12Wins["name"] == "Grand12Wins":
        Grand12WinsCout = Grand12Wins["progress"]

    for YearsPlayed in player_data["badges"]:
       if YearsPlayed["name"] == "YearsPlayed":
        YearsPlayedCout = YearsPlayed["progress"]

    for EmoteCollection in player_data["badges"]:
       if EmoteCollection["name"] == "EmoteCollection":
        EmoteCollectionCout = EmoteCollection["progress"]

    for BannerCollection in player_data["badges"]:
       if BannerCollection["name"] == "BannerCollection":
        BannerCollectionCout = BannerCollection["progress"]
    
    for TeamPlayer in player_data["achievements"]:
       if TeamPlayer["name"] == "Team Player":
        TeamPlayerCout = TeamPlayer["value"]

    for PracticewithFriendsCout in player_data["achievements"]:
      if PracticewithFriendsCout["name"] == "Practice with Friends":
        PracticewithFriendsCoutCout = PracticewithFriendsCout["value"]





    ws.append([
    player_data["name"],
    best2,best3,
    last2,last3,Classic12WinsCout,Grand12WinsCout,YearsPlayedCout,EmoteCollectionCout,BannerCollectionCout,TeamPlayerCout,PracticewithFriendsCoutCout,starPoints,totalExpPoints])

    x+=1
    print(x)
wb.save("TCRS.xlsx")

end_time = time.time()
print(f"執行時間：{end_time - start_time}")