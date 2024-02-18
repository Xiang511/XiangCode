import openpyxl
import requests
import time
from tqdm import tqdm

#計算起始時間
start_time = time.time()

# enter your key
# 前往 https://developer.clashroyale.com/#/ 取得

API_KEY = ""
headers = {
    "Authorization": "Bearer {}".format(API_KEY)
}

# enter your tag
# ex: '%2322R920J00','%2312345678'

player_tags = ['%2322R920J00']
              
wb = openpyxl.Workbook()
ws = wb.active

ws["A1"] = "遊戲ID"
ws["B1"] = "經典挑戰12勝次數"
ws["C1"] = "終極挑戰12勝次數"
ws["D1"] = "遊玩時間"
ws["E1"] = "表情收集次數"
ws["F1"] = "戰旗收集次數"
ws["G1"] = "星光點數"
ws["H1"] = "全部經驗值"


for player_tag in tqdm(player_tags):
    response = requests.get(f"https://api.clashroyale.com/v1/players/{player_tag}", headers=headers)
    
    player_data = response.json()

    Classic12WinsCout =0 
    Grand12WinsCout = 0

    starPoints = player_data["starPoints"]
    totalExpPoints = player_data["totalExpPoints"]
    
    for Classic12Wins in player_data["badges"]:
       if Classic12Wins["name"] == "Classic12Wins":
        Classic12WinsCout = Classic12Wins["progress"]

    for Grand12Wins in player_data["badges"]:
       if Grand12Wins["name"] == "Grand12Wins":
        Grand12WinsCout = Grand12Wins["progress"]

    for YearsPlayed in player_data["badges"]:
       if YearsPlayed["name"] == "YearsPlayed":
        YearsPlayedCout = YearsPlayed["progress"]

    for EmoteCollection in player_data["badges"]:
       if EmoteCollection["name"] == "EmoteCollection":
        EmoteCollectionCout = EmoteCollection["progress"]

    for BannerCollection in player_data["badges"]:
       if BannerCollection["name"] == "BannerCollection":
        BannerCollectionCout = BannerCollection["progress"]


    # for PracticewithFriendsCout in player_data["achievements"]:
    #   if PracticewithFriendsCout["name"] == "Practice with Friends":
    #     PracticewithFriendsCoutCout = PracticewithFriendsCout["value"]

    ws.append([
    player_data["name"],Classic12WinsCout,Grand12WinsCout,YearsPlayedCout,EmoteCollectionCout,BannerCollectionCout,starPoints,totalExpPoints
    ])

wb.save("個人成就.xlsx")

#計算結束時間
end_time = time.time()
print(f"執行時間：{end_time - start_time}")
