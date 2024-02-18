import requests
import openpyxl
import datetime
import time
from tqdm import tqdm

#計算起始時間
start_time = time.time()

# 設定 API 金鑰
# 前往 https://developer.clashroyale.com/#/ 取得

API_KEY = ""
headers = {
    "Authorization": "Bearer {}".format(API_KEY)
}

response = requests.get(
    "https://api.clashroyale.com/v1/locations/57000228/pathoflegend/players",
    headers=headers,
)

#   取得現在時間
now = datetime.datetime.now()
now_str = now.strftime("%Y-%m-%d")

wb = openpyxl.Workbook() 
ws = wb.active

ws.cell(row=1, column=1).value = "ID"
ws.cell(row=1, column=2).value = "Name"
ws.cell(row=1, column=3).value = "Rating"

row_number = 2
for player in tqdm(response.json()["items"]):
    ws.cell(row=row_number, column=1).value = player["tag"]
    ws.cell(row=row_number, column=2).value = player["name"]
    ws.cell(row=row_number, column=3).value = player["eloRating"]
    row_number += 1

wb.save(now_str+".xlsx")

#計算結束時間
end_time = time.time()
print(f"執行時間：{end_time - start_time}")
