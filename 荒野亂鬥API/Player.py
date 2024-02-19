import openpyxl
import requests
import json
import time
from tqdm import tqdm

# 計算起始時間
start_time = time.time()

# 設定 API 金鑰
# 前往 https://developer.clashroyale.com/#/ 取得

API_KEY = ""

headers = {
    "Authorization": "Bearer {}".format(API_KEY)
}

# enter your tag
# ex: '%2322R920J00','%2312345678'

player_tags = ["%2322CU2P2V","%23PL2LQJJU"]

wb = openpyxl.Workbook()
ws = wb.active

ws["A1"] = "tag"
ws["B1"] = "name"
ws["C1"] = "trophies"
ws["D1"] = "highestTrophies"
ws["E1"] = "expLevel"
ws["F1"] = "expPoints"
ws["G1"] = "Victories3v3"
ws["H1"] = "duoVictories"
ws["I1"] = "bestRoboRumbleTime"
ws["J1"] = "bestTimeAsBigBrawler"

def clean_result(result):
    if result is None:
        return ""
    return result[result.rfind(" ") + 1:].split("}")[0]

for player_tag in tqdm(player_tags):
    response = requests.get(f"https://api.brawlstars.com/v1/players/{player_tag}", headers=headers)
    
    player_data = response.json()
    trophies = json.dumps(player_data.get("trophies"))
    highestTrophies = json.dumps(player_data.get("highestTrophies"))
    expLevel = json.dumps(player_data.get("expLevel"))
    expPoints= json.dumps(player_data.get("expPoints"))
    Victories3v3= json.dumps(player_data.get("3vs3Victories"))
    soloVictories = json.dumps(player_data.get("soloVictories"))
    duoVictories= json.dumps(player_data.get("duoVictories"))
    bestRoboRumbleTime= json.dumps(player_data.get("bestRoboRumbleTime"))
    bestTimeAsBigBrawler= json.dumps(player_data.get("bestTimeAsBigBrawler"))
    ws.append(
        [
        player_data["tag"],
        player_data["name"],
        trophies,
        highestTrophies,
        highestTrophies,
        expLevel,
        expPoints,
        Victories3v3,
        duoVictories,
        bestRoboRumbleTime,
        bestTimeAsBigBrawler,
        ])
    
wb.save("Player.xlsx")

#計算結束時間
end_time = time.time()
print(f"執行時間：{end_time - start_time}")