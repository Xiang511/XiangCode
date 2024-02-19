import requests
import openpyxl
import datetime
import time


# 計算起始時間
start_time = time.time()

# 設定 API 金鑰
# 前往 https://developer.clashroyale.com/#/ 取得

API_KEY = ""
headers = {
    "Authorization": "Bearer {}".format(API_KEY)
}


response = requests.get(
    "https://api.brawlstars.com/v1/rankings/TW/players",
    headers=headers,
)


wb = openpyxl.Workbook() 

ws = wb.active
ws.cell(row=1, column=1).value = "tag"
ws.cell(row=1, column=2).value = "name"
ws.cell(row=1, column=2).value = "trophies"
ws.cell(row=1, column=2).value = "rank"

#   取得現在時間
now = datetime.datetime.now()
now_str = now.strftime("%Y-%m-%d %H:%M:%S")

ws["F1"] = "執行時間"
ws["F2"] = now_str


row_number = 2
for player in response.json()["items"]:
    # ws.cell(row=row_number, column=1).value = player["rank"]
    ws.cell(row=row_number, column=1).value = player["tag"]
    ws.cell(row=row_number, column=2).value = player["name"]
    ws.cell(row=row_number, column=3).value = player["trophies"]
    ws.cell(row=row_number, column=4).value = player["rank"]
    row_number += 1


wb.save(now_str+".xlsx")

#計算結束時間
end_time = time.time()
print(f"執行時間：{end_time - start_time}")